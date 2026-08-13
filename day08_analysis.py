from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,PatternFill

output_dir=Path("输出结果")
clean_data=pd.read_csv(
    output_dir/'clean_customer_usage.csv',
    encoding="utf-8-sig",
)

summary=(
    clean_data.groupby('region',as_index=False)
    .agg(
        customer_count=('customer_id','count'),
        total_usage=('usage','sum'),
        average_usage=('usage','mean'),
        total_amount=('amount','sum'),
        max_usage=('usage','max'),
    )
    .sort_values('total_amount',ascending=False)
)

summary[['total_usage','average_usage','total_amount','max_usage']]=summary[['total_usage','average_usage','total_amount','max_usage']].round(2)
output_file=output_dir/'day08_analysis.xlsx'
with pd.ExcelWriter(output_file,engine='openpyxl')as writer:
    clean_data.to_excel(writer,sheet_name='清洗明细',index=False)
    summary.to_excel(writer,sheet_name='区域汇总',index=False)

workbook=load_workbook(output_file)
for worksheet in workbook.worksheets:
    worksheet.freeze_panes='A2'

    for cell in worksheet[1]:
        cell.font=Font(bold=True,color='FFFFFF')
        cell.fill=PatternFill('solid',fgColor='1F4E79')
        cell.alignment=Alignment(horizontal='center')

    for column_cells in worksheet.columns:
        max_length=max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter=column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(
            max_length+3,
            24,
        )
workbook.save(output_file)
print(summary)
print("已生成：", output_file)
