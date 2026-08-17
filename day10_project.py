from pathlib import Path
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,PatternFill

plt.rcParams['font.sans-serif']=[
    "Microsoft YaHei",
    'SimHei',
    "Note Sans CJk SC",
]
plt.rcParams['axes.unicode_minus'] = False

output_dir=Path("输出结果")
output_dir.mkdir(exist_ok=True)
#读取Excel并且清洗数据
raw_data=pd.read_excel(
    '学习数据/training_data.xlsx',
    sheet_name='客户用量',
)
raw_data['customer_name']=raw_data['customer_name'].str.strip()
raw_data['region']=raw_data['region'].str.strip()
for column in ['previous_reading','current_reading','unit_price']:
    raw_data[column]=pd.to_numeric(
        raw_data[column],
        errors='coerce',
    )
raw_data['unit_price']=raw_data['unit_price'].fillna(0.62)
raw_data['usage']=(
    raw_data['current_reading']-raw_data['previous_reading']
)
#分离有效和异常数据
valid_mask=(
    raw_data['previous_reading'].notna()
    & raw_data['current_reading'].notna()
    & (raw_data['usage']>=0)
)
clean_data=raw_data.loc[valid_mask].copy()
error_data=raw_data.loc[~valid_mask].copy()
clean_data['amount']=(clean_data['unit_price']*clean_data['usage']).round(2)
#区域汇总
summary=clean_data.groupby('region',as_index=False).agg(
    customer_count=('customer_id','count'),
    total_usage=('usage','sum'),
    average_usage=('usage','mean'),
    total_amount=('amount','sum'),
    max_usage=('usage','max'),
)

summary[['total_usage','average_usage','total_amount','max_usage']]=summary[['total_usage','average_usage','total_amount','max_usage']].round(2)
summary=summary.sort_values('total_usage',ascending=False)

#写出并格式化综合 Excel 报告
report_file=output_dir/'final_data_analysis_report.xlsx'
with pd.ExcelWriter(report_file,engine='openpyxl') as writer:
    clean_data.to_excel(writer,sheet_name='有效明细',index=False)
    error_data.to_excel(writer,sheet_name='异常明细',index=False)
    summary.to_excel(writer,sheet_name='区域汇总',index=False)

workbook=load_workbook(report_file)
for worksheet in workbook.worksheets:
    worksheet.freeze_panes="A2"
    for cell in worksheet[1]:
        cell.font=Font(bold=True,color='FFFFFF')
        cell.fill=PatternFill('solid',fgColor='1F4E79')
        cell.alignment=Alignment(horizontal='center')

    for column_cells in worksheet.columns:
        max_length=max(
            len(str(cell.value))if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(max_length + 3, 24)

workbook.save(report_file)

#生成汇总柱状图并输出结果
plt.figure(figsize=[7,5])
bars=plt.bar(summary['region'],summary['total_usage'])
plt.bar_label(bars, fmt='%.1f', padding=1)
plt.title("综合练习：虚拟区域总用量")
plt.xlabel("区域")
plt.ylabel("总用量")
plt.tight_layout()

chart_file=output_dir / "final_region_usage.png"
plt.savefig(chart_file)
plt.close()

print("有效记录：", len(clean_data))
print("异常记录：", len(error_data))
print(summary)
print("已生成：", report_file)
print("已生成：", chart_file)


